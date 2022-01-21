import requests
from .models import (GetTokenML,DocumentItems,DictionaryItems)
import environ
env = environ.Env()
environ.Env.read_env()
import json
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from datetime import datetime,timedelta
from django.utils import timezone
from django.core.files import File
import os
import pandas
import string

listTitlesExcel = (
        "Id",
        "Categoría",
        "Título",
        "Descripción",
        "Precio",
        "SKU",
        "Estado",
        "Stock",
        "Tipo de Publicación",
        "Condición",
        "Envio Gratis",
        "Precio Envio Gratis",
        "Modo Envio",
        "Metodo Envio",
        "Retira en Persona",
        "Garantia",
        "Fecha Creación",
        "Última Actualización",
        "Resultado",
        "Resultado Observaciones",
        "Imagen 1",
        "Imagen 2",
        "Imagen 3",
        "Imagen 4",
        "Imagen 5",
        "Imagen 6",
        "Imagen 7",
        "Imagen 8",
        "Imagen 9",
        "Imagen 10",
        "Video",
        "Calidad de la Publicación",
        "Calidad de la Imagen",
        "Mejoras pendientes",
        "Atributo \r\n Marca",
        "Atributo \r\n Modelo",
        "Atributo \r\n Número de parte",
        "Atributo \r\n Cantidad de leds por lámpara",
        "Atributo \r\n Incluye balastros",
        "Atributo \r\n Temperatura",
        "Atributo \r\n Consumo",
        "Atributo \r\n Código universal de producto",
        "Atributo \r\n OEM",
        "Atributo \r\n MPN",
        "Atributo \r\n Material",
        "Atributo \r\n Tipo de buje de control de suspensión",
        "Atributo \r\n Color de la luz",
        "Atributo \r\n Tipo de conector",
        "Atributo \r\n Color del bulbo",
        "Atributo \r\n Tecnología del bulbo",
        "Atributo \r\n Cantidad de bulbos",
        "Atributo \r\n Color principal",
        "Atributo \r\n Material del aislante",
        "Atributo \r\n Diámetro exterior del aislante",
        "Atributo \r\n Es resistivo",
        "Atributo \r\n Material del núcleo",
        "Atributo \r\n Posición",
        "Atributo \r\n Incluye accesorios para el montaje",
        "Atributo \r\n Unidades por envase",
        "Atributo \r\n Peso neto",
        "Atributo \r\n Largo",
        "Atributo \r\n Diámetro",
        "Atributo \r\n Origen",
        "Atributo \r\n Impedancia",
        "Atributo \r\n Tipo de pintura",
        "Atributo \r\n Volumen neto",
        "Atributo \r\n Formato de venta",
        "Atributo \r\n Superficies aptas",
        "Atributo \r\n Tipo de base",
        "Atributo \r\n Volumen de la unidad",
        "Atributo \r\n Tipo de pegamento",
        "Atributo \r\n Tecnología de iluminación",
        "Atributo \r\n Watts",
        "Atributo \r\n Voltaje",
        "Atributo \r\n Tipos de rosca",
        "Atributo \r\n Forma",
        "Atributo \r\n Tamaño",
        "Atributo \r\n Diámetro de la boca",
        "Atributo \r\n Largo total",
        "Atributo \r\n Lado",
        "Atributo \r\n Cantidad de balatas",
        "Atributo \r\n Incluye sensores de desgaste",
        "Atributo \r\n Código FMSI",
        "Atributo \r\n Fabricante",
        "Atributo \r\n Unidades por pack",
        "Atributo \r\n Cantidad de bujías",
        "Atributo \r\n Presentación",
        "Atributo \r\n Fragancia",
        "Atributo \r\n Formato del desinfectante y limpiador multiuso",
        "Atributo \r\n Superficies recomendadas",
        "Atributo \r\n Áreas de limpieza recomendadas",
        "Atributo \r\n Es producto desinfectante",
        "Atributo \r\n Es inflamable",
        "Atributo \r\n Ancho",
        "Atributo \r\n Cantidad de estrías lado rueda",
        "Atributo \r\n Cantidad de estrías lado caja",
        "Atributo \r\n Volumen",
        "Atributo \r\n Diámetro máximo de sellado",
        "Atributo \r\n Tipo de llanta",
        "Atributo \r\n Tipo de envase",
        "Atributo \r\n Diámetro alambre",
        "Atributo \r\n Eje",
        "Atributo \r\n Incluye bujes",
        "Atributo \r\n Tipo de inyección",
        "Atributo \r\n Incluye válvula IAC",
        "Atributo \r\n Incluye sensor de posición",
        "Atributo \r\n Incluye empaque",
        "Atributo \r\n Incluye tornillos de montaje",
        "Atributo \r\n Posiciones",
        "Atributo \r\n Incluye bases",
        "Atributo \r\n Incluye topes",
        "Atributo \r\n Incluye resortes",
        "Atributo \r\n Posición del eje",
        "Atributo \r\n Posición de la horquilla",
        "Atributo \r\n Incluye rótula",
        "Atributo \r\n Incluye buje",
        "Atributo \r\n Incluye grasa",
        "Atributo \r\n Es pre engrasada",
        "Atributo \r\n Incluye kit de montaje",
        "Atributo \r\n Es reemplazo original",
        "Atributo \r\n Línea",
        "Atributo \r\n Corriente de salida",
        "Atributo \r\n Tensión de carga",
        "Atributo \r\n Potencia máxima",
        "Atributo \r\n Color",
        "Atributo \r\n Unidad de venta",
        "Atributo \r\n Rendimiento",
        "Atributo \r\n Ambientes",
        "Atributo \r\n Tiempo de secado",
        "Atributo \r\n Es aerosol",
        "Atributo \r\n Carcasa del diferencial incluída",
        "Atributo \r\n Tipo de uso",
        "Atributo \r\n Capacidad de fluído del diferencial",
        "Atributo \r\n Relación de transmisión del diferencial",
        "Atributo \r\n Largo de la aguja",
        "Atributo \r\n Capacidad en volumen",
        "Atributo \r\n Material de la jeringa",
        "Atributo \r\n Material de la aguja",
        "Atributo \r\n Incluye agujas extras",
        "Atributo \r\n Incluye filtro de aire",
        "Atributo \r\n Incluye filtro de aceite",
        "Atributo \r\n Incluye filtro de aire acondicionado",
        "Atributo \r\n Incluye filtro de combustible",
        "Atributo \r\n Incluye aceite",
        "Atributo \r\n Tipo de filtro de aceite",
        "Atributo \r\n Formato",
        "Atributo \r\n Es apto para lavarropas",
        "Atributo \r\n Es blanqueador",
        "Atributo \r\n Tipo de funcionamiento de la válvula EGR",
        "Atributo \r\n Género del terminal",
        "Atributo \r\n Diámetro de los puertos de vacío",
        "Atributo \r\n Cantidad de puertos de vacío",
        "Atributo \r\n Cantidad de terminales",
        "Atributo \r\n Cantidad de huecos de montaje de la válvula EGR",
        "Atributo \r\n Juntas incluidas	",
        "Atributo \r\n Posición de la rótula de suspensión",
        "Atributo \r\n Diámetro externo de la rótula de suspensión",
        "Atributo \r\n Diámetro de la rosca del tornillo",
        "Atributo \r\n Tuerca incluida",
        "Atributo \r\n Diámetro de entrada",
        "Atributo \r\n Diámetro de salida",
        "Atributo \r\n Número de DOT",
        "Atributo \r\n Grado del aceite de motor",
        "Atributo \r\n Tipo de aceite de motor",
        "Atributo \r\n Categoría de servicio",
        "Atributo \r\n Volumen del aceite de motor",
        "Atributo \r\n Tipo de contenedor",
        "Atributo \r\n Ubicación",
        "Atributo \r\n Tipo de luz",
        "Atributo \r\n Peso de la unidad",
        "Atributo \r\n Formato del abrillantador",
        "Atributo \r\n Cantidad de unidades",
        "Atributo \r\n Contenido neto",
        "Atributo \r\n Set"
        )

listRowExcel = (
    "Id",
    "Categoría",
    "Título",
    "Descripción",
    "Precio",
    "SKU",
    "Estado",
    "Stock",
    "Tipo de Publicación",
    "Condición",
    "Envio Gratis",
    "Precio Envio Gratis",
    "Modo Envio",
    "Metodo Envio",
    "Retira en Persona",
    "Garantia",
    "Fecha Creación",
    "Última Actualización",
    "Resultado",
    "Resultado Observaciones",
    "Imagen 1",
    "Imagen 2",
    "Imagen 3",
    "Imagen 4",
    "Imagen 5",
    "Imagen 6",
    "Imagen 7",
    "Imagen 8",
    "Imagen 9",
    "Imagen 10",
    "Video",
    "Calidad de la Publicación",
    "Calidad de la Imagen",
    "Mejoras pendientes",
    "Marca",
    "Modelo",
    "Número de parte",
    "Cantidad de leds por lámpara",
    "Incluye balastros",
    "Temperatura",
    "Consumo",
    "Código universal de producto",
    "OEM",
    "MPN",
    "Material",
    "Tipo de buje de control de suspensión",
    "Color de la luz",
    "Tipo de conector",
    "Color del bulbo",
    "Tecnología del bulbo",
    "Cantidad de bulbos",
    "Color principal",
    "Material del aislante",
    "Diámetro exterior del aislante",
    "Es resistivo",
    "Material del núcleo",
    "Posición",
    "Incluye accesorios para el montaje",
    "Unidades por envase",
    "Peso neto",
    "Largo",
    "Diámetro",
    "Origen",
    "Impedancia",
    "Tipo de pintura",
    "Volumen neto",
    "Formato de venta",
    "Superficies aptas",
    "Tipo de base",
    "Volumen de la unidad",
    "Tipo de pegamento",
    "Tecnología de iluminación",
    "Watts",
    "Voltaje",
    "Tipos de rosca",
    "Forma",
    "Tamaño",
    "Diámetro de la boca",
    "Largo total",
    "Lado",
    "Cantidad de balatas",
    "Incluye sensores de desgaste",
    "Código FMSI",
    "Fabricante",
    "Unidades por pack",
    "Cantidad de bujías",
    "Presentación",
    "Fragancia",
    "Formato del desinfectante y limpiador multiuso",
    "Superficies recomendadas",
    "Áreas de limpieza recomendadas",
    "Es producto desinfectante",
    "Es inflamable",
    "Ancho",
    "Cantidad de estrías lado rueda",
    "Cantidad de estrías lado caja",
    "Volumen",
    "Diámetro máximo de sellado",
    "Tipo de llanta",
    "Tipo de envase",
    "Diámetro alambre",
    "Eje",
    "Incluye bujes",
    "Tipo de inyección",
    "Incluye válvula IAC",
    "Incluye sensor de posición",
    "Incluye empaque",
    "Incluye tornillos de montaje",
    "Posiciones",
    "Incluye bases",
    "Incluye topes",
    "Incluye resortes",
    "Posición del eje",
    "Posición de la horquilla",
    "Incluye rótula",
    "Incluye buje",
    "Incluye grasa",
    "Es pre engrasada",
    "Incluye kit de montaje",
    "Es reemplazo original",
    "Línea",
    "Corriente de salida",
    "Tensión de carga",
    "Potencia máxima",
    "Color",
    "Unidad de venta",
    "Rendimiento",
    "Ambientes",
    "Tiempo de secado",
    "Es aerosol",
    "Carcasa del diferencial incluída",
    "Tipo de uso",
    "Capacidad de fluído del diferencial",
    "Relación de transmisión del diferencial",
    "Largo de la aguja",
    "Capacidad en volumen",
    "Material de la jeringa",
    "Material de la aguja",
    "Incluye agujas extras",
    "Incluye filtro de aire",
    "Incluye filtro de aceite",
    "Incluye filtro de aire acondicionado",
    "Incluye filtro de combustible",
    "Incluye aceite",
    "Tipo de filtro de aceite",
    "Formato",
    "Es apto para lavarropas",
    "Es blanqueador",
    "Tipo de funcionamiento de la válvula EGR",
    "Género del terminal",
    "Diámetro de los puertos de vacío",
    "Cantidad de puertos de vacío",
    "Cantidad de terminales",
    "Cantidad de huecos de montaje de la válvula EGR",
    "Juntas incluidas	",
    "Posición de la rótula de suspensión",
    "Diámetro externo de la rótula de suspensión",
    "Diámetro de la rosca del tornillo",
    "Tuerca incluida",
    "Diámetro de entrada",
    "Diámetro de salida",
    "Número de DOT",
    "Grado del aceite de motor",
    "Tipo de aceite de motor",
    "Categoría de servicio",
    "Volumen del aceite de motor",
    "Tipo de contenedor",
    "Ubicación",
    "Tipo de luz",
    "Peso de la unidad",
    "Formato del abrillantador",
    "Cantidad de unidades",
    "Contenido neto",
    "Set"
    )

def UpdateItem(Originalitem,newData):
        Originalitem.stock = newData['instk']
        Originalitem.save()
        # try:
        #     tokenSave = GetTokenML()
        #     #make a format to send to mercado libre API
        #     data = {}
        #     quantity = newData['instk']
        #     #if the price and quantity are none, ignore
        #     if quantity is not None:
        #         data["available_quantity"]= quantity
        #     item = Originalitem.idMercadoLibre
        #     #send the request to mercado libre and show the result
        #     headers = {'Authorization': 'Bearer '+tokenSave.access_token}
        #     url = env("APIURLML")+'items/'+str(item)
        #     response = requests.put(url, headers=headers,data=json.dumps(data))
        #     responsejson = response.json()
        #     print(responsejson)
        #     if 'id' in responsejson:
        #         print("Se actualizo")
        #     else:
        #         print("No se pudo actualizar")
        # except Exception as excep:
        #     print("No se pudo actualizar")

def getItemFromMLAPI(item,newData):
    try:
        tokenSave = GetTokenML()
        headers = {'Authorization': 'Bearer '+tokenSave.access_token}
        url = env("APIURLML")+'items?ids='+str(item.idMercadoLibre)
        response = requests.get(url, headers=headers)
        responsejson = response.json()
        dataCleanResponse = responsejson[0]["body"]
        dataitem = {}
        #Id
        dataitem["Id"] = dataCleanResponse["id"]
        #Categoría
        urlcategory = env("APIURLML")+'categories/'+str(dataCleanResponse["category_id"])
        responseCategory = requests.get(urlcategory)
        responseCategoryJson = responseCategory.json()
        categoryString = ""
        for category in responseCategoryJson['path_from_root']:
            categoryString += str(category['name']) + " > "
        categoryString += str(responseCategoryJson['name'])
        dataitem['Categoría'] = categoryString
        #Titulo
        dataitem['Título'] = dataCleanResponse["title"]
        #Descripción
        urlDescription = env("APIURLML")+'items/'+str(dataCleanResponse["id"])+'/description'
        responseDescription = requests.get(urlDescription, headers=headers)
        responseDescriptionJson = responseDescription.json()
        dataitem['Descripción'] = str(responseDescriptionJson.get('plain_text',"")).replace('\n','\r\n')
        #Precio
        dataitem['Precio'] = newData['cost']
        #Obtener Atributos
        atrributes = dataCleanResponse["attributes"]
        #SKU
        sku = next((item for item in atrributes if item["name"] == "SKU"), None)
        if sku is not None:
            dataitem['SKU'] = sku['value_name']
        #Estado
        if dataCleanResponse['status'] == 'active':
            espstatus = "Activa"
        elif dataCleanResponse['status'] == 'paused':
            espstatus = "Pausada"
        dataitem['Estado'] = espstatus
        #Stock
        dataitem['Stock'] = newData['instk']
        #Tipo de Publicación
        listing_type = {
            "gold_pro" : "Premium",
            "gold_premium":"Oro Premium",
            "gold_special":"Clásica",
            "gold":"Oro",
            "silver":"Plata",
            "bronze":"Bronce",
            "free":"Gratuita"
        }
        dataitem['Tipo de Publicación'] = listing_type.get(dataCleanResponse['listing_type_id'],"No registrada")
        #Condición
        condition = next((item for item in atrributes if item["id"] == "ITEM_CONDITION"), None)
        if condition is not None:
            dataitem['Condición'] = condition['value_name']
        #Envio Gratis
        shipping = ""
        if dataCleanResponse['shipping']['free_shipping'] == True:
            shipping = "Si"
        else:
            shipping = "No"
        dataitem['Envio Gratis'] = shipping
        #Precio Envio Gratis
        if shipping == "Si":
            urlShipment = env("APIURLML")+'items/'+str(dataCleanResponse["id"])+'/shipping_options?zip_code=22454'
            responseDescription = requests.get(urlShipment, headers=headers)
            responseDescriptionJson = responseDescription.json()
            try:
                optionShiping = next((item for item in responseDescriptionJson['options'] if item["name"] == "Estándar a domicilio"), None)
                if optionShiping is not None:
                    dataitem['Precio Envio Gratis'] = optionShiping['list_cost']
            except Exception as excep:
                dataitem['Precio Envio Gratis'] = ""
        else:
            dataitem['Precio Envio Gratis'] = ""
        #Modo Envio
        if dataCleanResponse['shipping']['mode'] == "custom":
            dataitem['Modo Envio'] = "Personalizado"
        elif dataCleanResponse['shipping']['mode'] == "me2":
            dataitem['Modo Envio'] = "MercadoEnvios2"
        else:
            dataitem['Modo Envio'] = "No Especificado"
        #Metodo Envio
        #Envio Gratis Si
        if (dataitem['Envio Gratis'] == "Si" and dataitem['Modo Envio'] == "MercadoEnvios2"):
            dataitem['Metodo Envio'] = "Estándar a domicilio"
        elif (dataitem['Envio Gratis'] == "Si" and dataitem['Modo Envio'] == "No Especificado"):
            dataitem['Metodo Envio'] = ""
        #Envio Gratis no
        elif (dataitem['Envio Gratis'] == "No" and (dataitem['Modo Envio'] == "MercadoEnvios2" or dataitem['Modo Envio'] == "No Especificado")):
            dataitem['Metodo Envio'] = ""
        elif (dataitem['Envio Gratis'] == "No" and dataitem['Modo Envio'] == "Personalizado"):
            urlShipment = env("APIURLML")+'items/'+str(dataCleanResponse["id"])+'/shipping_options?zip_code=22454'
            responseDescription = requests.get(urlShipment, headers=headers)
            responseDescriptionJson = responseDescription.json()
            options = ""
            for item in responseDescriptionJson['options']:
                options += item['name'] +": $"+str(item['cost']) + " \r\n"
            dataitem['Metodo Envio'] = options
        else:
            dataitem['Metodo Envio'] = ""
        #Retira en persona
        if dataCleanResponse['shipping']['local_pick_up'] == True:
            dataitem['Retira en Persona'] = "Si"
        else:
            dataitem['Retira en Persona'] = "No"
        #Garantia
        dataitem['Garantia'] = dataCleanResponse['warranty']
        #Fecha Creación
        dateCreatedConvert = datetime.strptime(dataCleanResponse['date_created'], '%Y-%m-%dT%H:%M:%S.%fZ')
        dateCreatedAdjust = dateCreatedConvert + timedelta(hours=-7)
        dataitem['Fecha Creación'] = dateCreatedAdjust.strftime("%d/%m/%Y %H:%M:%S%p")
        #Última Actualización
        dateUpdatedConvert = datetime.strptime(dataCleanResponse['last_updated'], '%Y-%m-%dT%H:%M:%S.%fZ')
        dateUpdatedAdjust = dateUpdatedConvert + timedelta(hours=-7)
        dataitem['Última Actualización'] = dateUpdatedAdjust.strftime("%d/%m/%Y %H:%M:%S%p")
        #Imagenes
        contpicture = 1 
        for picture in dataCleanResponse['pictures']:
            dataitem['Imagen '+str(contpicture)] = picture['secure_url']
            contpicture += 1
        #Video
        videoid = dataCleanResponse['video_id']
        if videoid is not None:
            dataitem['Video'] = "https://www.youtube.com/watch?v="+videoid
        else:
            dataitem['Video'] = ""
        #Calidad de la publicación
        qualityMessages = env("APIURLML")+'items/'+str(dataCleanResponse["id"])+('/health/actions')
        responseQuality = requests.get(qualityMessages,headers=headers)
        responseQualityJson = responseQuality.json()
        print(responseQualityJson)
        health = float(responseQualityJson.get("health",-1))
        healthText = ""
        print(health)
        if health == -1:
            healthText = "Desconocido"
        else:
            healthText = str(health * 100) + "%"
        dataitem['Calidad de la Publicación'] = healthText
        #Calidad de la imagen
        #Mejoras pendientes
        acctions = {
            "technical_specification":"verifica la calidad de los atributos y completa la ficha técnica.",
            "buybox":"publica en catálogo.",
            "variations":"utiliza variaciones para la publicación.",
            "product_identifiers": "informar código universal del producto.",
            "picture": "verifica la calidad de las imágenes.",
            "price": "publica con precio más competitivo.",
            "me2": "utiliza Mercado Envíos en las publicaciones.",
            "free_shipping":"ofrece envíos gratis.",
            "flex": "utiliza Mercado Envíos Flex.",
            "immediate_payment":"utiliza Mercado Pago.",
            "classic": "realiza una publicación con exposición al menos clásica.",
            "premium": "realiza una publicación como premium.",
            "size_chart": "informa una guía de tallas.",
            "publish": "relacionado a la publicación del ítem"
        }
        publishactions = responseQualityJson.get("actions",[])
        textActions = ""
        if publishactions is not None:
            for act in publishactions:
                if act['id'] == "picture":
                    dataitem['Calidad de la Imagen'] = "verifica la calidad de las imágenes."
                else:
                    textActions += acctions.get(act['id'],"Desconocido") + "\r\n"
        dataitem['Mejoras pendientes'] = textActions
        #Marca
        data = next((item for item in atrributes if item["id"] == "BRAND"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Marca'] = data["value_name"]
        #Modelo
        data = next((item for item in atrributes if item["id"] == "MODEL"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Modelo'] = data["value_name"]
        #Número de parte
        data = next((item for item in atrributes if item["id"] == "PART_NUMBER"), None)
        if data is not None:
            if data["value_name"] is not None:
                dataitem['Número de parte'] = data["value_name"]
        #Cantidad de leds por lámpara
        data = next((item for item in atrributes if item["id"] == "LED_LIGHTS_NUMBER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de leds por lámpara'] = data["value_name"]
        #Incluye balastros
        data = next((item for item in atrributes if item["id"] == "INCLUDES_BALLASTS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye balastros'] = data["value_name"]
        #Temperatura
        data = next((item for item in atrributes if item["id"] == "CAR_LED_BULB_TEMPERATURE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Temperatura'] = data["value_name"]
        #Consumo
        data = next((item for item in atrributes if item["id"] == "CONSUMPTION"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Consumo'] = data["value_name"]
        #Código universal de producto
        data = next((item for item in atrributes if item["id"] == "GTIN"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Código universal de producto'] = data["value_name"]
        #OEM
        data = next((item for item in atrributes if item["id"] == "OEM"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['OEM'] = data["value_name"]
        #Material
        data = next((item for item in atrributes if item["id"] == "MATERIAL"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Material'] = data["value_name"]
        #Material
        data = next((item for item in atrributes if item["id"] == "MATERIAL"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Material'] = data["value_name"]
        #Tipo de buje de control de suspensión
        data = next((item for item in atrributes if item["id"] == "SUSPENSION_CONTROL_ARM_BUSHING_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de buje de control de suspensión'] = data["value_name"]
        #Color de la luz
        data = next((item for item in atrributes if item["id"] == "LIGHT_COLOR"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Color de la luz'] = data["value_name"]
        #Tipo de conector
        data = next((item for item in atrributes if item["id"] == "CAR_LIGHT_BULB_CONNECTOR_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de conector'] = data["value_name"]
        #Color del bulbo
        data = next((item for item in atrributes if item["id"] == "BULB_COLOR"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Color del bulbo'] = data["value_name"]
        #Tecnología del bulbo
        data = next((item for item in atrributes if item["id"] == "BULB_TECHNOLOGY"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tecnología del bulbo'] = data["value_name"]
        #Cantidad de bulbos
        data = next((item for item in atrributes if item["id"] == "BULBS_NUMBER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de bulbos'] = data["value_name"]
        #Color principal
        data = next((item for item in atrributes if item["id"] == "MAIN_COLOR"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Color principal'] = data["value_name"]
        #Material del aislante
        data = next((item for item in atrributes if item["id"] == "INSULATION_MATERIAL"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Material del aislante'] = data["value_name"]
        #Diámetro exterior del aislante
        data = next((item for item in atrributes if item["id"] == "INSULATION_OUTSIDE_DIAMETER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Diámetro exterior del aislante'] = data["value_name"]
        #Es resistivo
        data = next((item for item in atrributes if item["id"] == "IS_RESISTIVE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Es resistivo'] = data["value_name"]
        #Material del núcleo
        data = next((item for item in atrributes if item["id"] == "CORE_MATERIAL"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Material del núcleo'] = data["value_name"]
        #Material del núcleo
        data = next((item for item in atrributes if item["id"] == "CORE_MATERIAL"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Material del núcleo'] = data["value_name"]
        #Posición
        data = next((item for item in atrributes if item["id"] == "POSITION"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Posición'] = data["value_name"]
        #Incluye accesorios para el montaje
        data = next((item for item in atrributes if item["id"] == "INCLUDES_MOUNTING_ACCESSORIES"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye accesorios para el montaje'] = data["value_name"]
        #Unidades por envase
        data = next((item for item in atrributes if item["id"] == "UNITS_PER_PACKAGE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Unidades por envase'] = data["value_name"]
        #Peso neto
        data = next((item for item in atrributes if item["id"] == "NET_WEIGHT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Peso neto'] = data["value_name"]
        #Largo
        data = next((item for item in atrributes if item["id"] == "LENGTH"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Largo'] = data["value_name"]
        #Diámetro
        data = next((item for item in atrributes if item["id"] == "DIAMETER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Diámetro'] = data["value_name"]
        #Origen
        data = next((item for item in atrributes if item["id"] == "ORIGIN"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Origen'] = data["value_name"]
        #Tipo de pintura
        data = next((item for item in atrributes if item["id"] == "PAINT_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de pintura'] = data["value_name"]
        #Volumen neto
        data = next((item for item in atrributes if item["id"] == "NET_VOLUME"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Volumen neto'] = data["value_name"]
        #Formato de venta
        data = next((item for item in atrributes if item["id"] == "SALE_FORMAT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Formato de venta'] = data["value_name"]
        #Superficies aptas
        data = next((item for item in atrributes if item["id"] == "APT_SURFACES"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Superficies aptas'] = data["value_name"]
        #Tipo de base
        data = next((item for item in atrributes if item["id"] == "BASE_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de base'] = data["value_name"]
        #Volumen de la unidad
        data = next((item for item in atrributes if item["id"] == "UNIT_VOLUME"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Volumen de la unidad'] = data["value_name"]
        #Tipo de pegamento
        data = next((item for item in atrributes if item["id"] == "GLUE_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de pegamento'] = data["value_name"]
        #Tecnología de iluminación
        data = next((item for item in atrributes if item["id"] == "LIGHTING_TECHNOLOGY"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tecnología de iluminación'] = data["value_name"]
        #Watts
        data = next((item for item in atrributes if item["id"] == "WATTAGE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Watts'] = data["value_name"]
        #Voltaje
        data = next((item for item in atrributes if item["id"] == "VOLTAGE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Voltaje'] = data["value_name"]
        #Tipos de rosca 
        data = next((item for item in atrributes if item["id"] == "FITTING_TYPES"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipos de rosca'] = data["value_name"]
        #Forma
        data = next((item for item in atrributes if item["id"] == "LIGHT_BULB_SHAPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Forma'] = data["value_name"]
        #Tamaño
        data = next((item for item in atrributes if item["id"] == "FUNNEL_SIZE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tamaño'] = data["value_name"]
        #Diámetro de la boca
        data = next((item for item in atrributes if item["id"] == "MOUTH_DIAMETER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Diámetro de la boca'] = data["value_name"]
        #Diámetro de la boca
        data = next((item for item in atrributes if item["id"] == "MOUTH_DIAMETER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Diámetro de la boca'] = data["value_name"]
        #Largo total
        data = next((item for item in atrributes if item["id"] == "TOTAL_LENGTH"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Largo total'] = data["value_name"]
        #Lado
        data = next((item for item in atrributes if (item["id"] == "SIDE_POSITION" or item["id"] == "SIDE")), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Lado'] = data["value_name"]
        #Cantidad de balatas
        data = next((item for item in atrributes if item["id"] == "QUANTITY_OF_PADS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de balatas'] = data["value_name"]
        #Incluye sensores de desgaste
        data = next((item for item in atrributes if item["id"] == "INCLUDES_WEAR_SENSORS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye sensores de desgaste'] = data["value_name"]
        #Código FMSI
        data = next((item for item in atrributes if item["id"] == "FMSI_NUMBER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Código FMSI'] = data["value_name"]
        #Fabricante
        data = next((item for item in atrributes if item["id"] == "MANUFACTURER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Fabricante'] = data["value_name"]
        #Unidades por pack
        data = next((item for item in atrributes if item["id"] == "UNITS_PER_PACK"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Unidades por pack'] = data["value_name"]
        #Cantidad de bujías
        data = next((item for item in atrributes if item["id"] == "NUMBER_OF_SPARK_PLUGS_BY_KIT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de bujías'] = data["value_name"]
        #Presentación
        data = next((item for item in atrributes if item["id"] == "PRESENTATION"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Presentación'] = data["value_name"]
        #Fragancia
        data = next((item for item in atrributes if item["id"] == "FRAGRANCE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Fragancia'] = data["value_name"]
        #Formato del desinfectante y limpiador multiuso
        data = next((item for item in atrributes if item["id"] == "MULTIPURPOSE_CLEANER_AND_DISINFECTANT_FORMAT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Formato del desinfectante y limpiador multiuso'] = data["value_name"] 
        #Superficies recomendadas
        data = next((item for item in atrributes if item["id"] == "RECOMMENDED_SURFACES"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Superficies recomendadas'] = data["value_name"] 
        #Áreas de limpieza recomendadas
        data = next((item for item in atrributes if item["id"] == "RECOMMENDED_CLEANING_AREAS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Áreas de limpieza recomendadas'] = data["value_name"] 
        #Es producto desinfectante
        data = next((item for item in atrributes if item["id"] == "IS_DISINFECTANT_PRODUCT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Es producto desinfectante'] = data["value_name"] 
        #Es inflamable
        data = next((item for item in atrributes if item["id"] == "IS_FLAMMABLE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Es inflamable'] = data["value_name"] 
        #Ancho
        data = next((item for item in atrributes if item["id"] == "WIDTH"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Ancho'] = data["value_name"] 
        #Cantidad de estrías lado rueda
        data = next((item for item in atrributes if item["id"] == "NUMBER_TEETH_WHEEL_SIDE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de estrías lado rueda'] = data["value_name"] 
        #Cantidad de estrías lado caja
        data = next((item for item in atrributes if item["id"] == "NUMBER_TEETH_HUB_SIDE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de estrías lado caja'] = data["value_name"] 
        #Volumen
        data = next((item for item in atrributes if item["id"] == "VOLUME_SIZE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Volumen'] = data["value_name"]
        #Diámetro máximo de sellado
        data = next((item for item in atrributes if item["id"] == "MAX_SEALING_DIAMETER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Diámetro máximo de sellado'] = data["value_name"]
        #Tipo de llanta
        data = next((item for item in atrributes if item["id"] == "TIRE_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de llanta'] = data["value_name"]
        #Tipo de envase
        data = next((item for item in atrributes if item["id"] == "PACKAGING_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de envase'] = data["value_name"]
        #Diámetro alambre
        data = next((item for item in atrributes if item["id"] == "BAR_DIAMETER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Diámetro alambre'] = data["value_name"]
        #Incluye bujes
        data = next((item for item in atrributes if item["id"] == "BUSHING_INCLUDED"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye bujes'] = data["value_name"]
        #Tipo de inyección
        data = next((item for item in atrributes if item["id"] == "INJECTION_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de inyección'] = data["value_name"]
        #Incluye válvula IAC
        data = next((item for item in atrributes if item["id"] == "INCLUDES_IDLE_AIR_CONTROL_VALVE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye válvula IAC'] = data["value_name"] 
        #Incluye sensor de posición
        data = next((item for item in atrributes if item["id"] == "INCLUDES_POSITION_SENSOR"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye sensor de posición'] = data["value_name"]
        #Incluye empaque
        data = next((item for item in atrributes if item["id"] == "INCLUDES_GASKET"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye empaque'] = data["value_name"]
        #Incluye tornillos de montaje
        data = next((item for item in atrributes if item["id"] == "INCLUDES_MOUNTING_SCREWS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye tornillos de montaje'] = data["value_name"]
        #Posiciones
        data = next((item for item in atrributes if item["id"] == "POSITIONS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Posiciones'] = data["value_name"]
        #Incluye bases
        data = next((item for item in atrributes if item["id"] == "INCLUDES_STRUT_MOUNTS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye bases'] = data["value_name"]
        #Incluye topes
        data = next((item for item in atrributes if item["id"] == "INCLUDES_BUMP_STOPS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye topes'] = data["value_name"]
        #Incluye resortes
        data = next((item for item in atrributes if item["id"] == "INCLUDES_SPRINGS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye resortes'] = data["value_name"]
        #Posición del eje
        data = next((item for item in atrributes if item["id"] == "AXIS_POSITION"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Posición del eje'] = data["value_name"]
        #Posición de la horquilla
        data = next((item for item in atrributes if item["id"] == "CONTROL_ARM_POSITION"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Posición de la horquilla'] = data["value_name"]
        #Incluye rótula
        data = next((item for item in atrributes if item["id"] == "INCLUDES_BALL_JOINT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye rótula'] = data["value_name"]
        #Incluye buje
        data = next((item for item in atrributes if item["id"] == "INCLUDES_BUSHING"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye buje'] = data["value_name"]
        #Incluye grasa
        data = next((item for item in atrributes if item["id"] == "INCLUDES_GREASE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye grasa'] = data["value_name"]
        #Es pre engrasada
        data = next((item for item in atrributes if item["id"] == "IS_PRE_GREASED"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Es pre engrasada'] = data["value_name"]
        #Incluye kit de montaje
        data = next((item for item in atrributes if item["id"] == "INCLUDES_MOUNTING_HARDWARE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye kit de montaje'] = data["value_name"]
        #Es reemplazo original
        data = next((item for item in atrributes if item["id"] == "IS_OEM_REPLACEMENT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Es reemplazo original'] = data["value_name"]
        #Línea
        data = next((item for item in atrributes if item["id"] == "LINE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Línea'] = data["value_name"]
        #Corriente de salida
        data = next((item for item in atrributes if item["id"] == "OUTPUT_CURRENT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Corriente de salida'] = data["value_name"]
        #Tensión de carga
        data = next((item for item in atrributes if item["id"] == "LOAD_VOLTAGE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tensión de carga'] = data["value_name"]
        #Potencia máxima
        data = next((item for item in atrributes if item["id"] == "ELECTRICAL_MAXIMUM_POWER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Potencia máxima'] = data["value_name"]
        #Color
        data = next((item for item in atrributes if item["id"] == "COLOR"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Color'] = data["value_name"]
        #Unidad de venta
        data = next((item for item in atrributes if item["id"] == "SALES_UNIT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Unidad de venta'] = data["value_name"]
        #Ambientes
        data = next((item for item in atrributes if item["id"] == "AMBIENTS"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Ambientes'] = data["value_name"]
        #Tiempo de secado
        data = next((item for item in atrributes if item["id"] == "DRY_TIME"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tiempo de secado'] = data["value_name"]
        #Es aerosol
        data = next((item for item in atrributes if item["id"] == "IS_AEROSOL"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Es aerosol'] = data["value_name"]
        #Incluye filtro de aire
        data = next((item for item in atrributes if item["id"] == "INCLUDES_AIR_FILTER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye filtro de aire'] = data["value_name"]
        #Incluye filtro de aceite
        data = next((item for item in atrributes if item["id"] == "INCLUDES_OIL_FILTER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye filtro de aceite'] = data["value_name"]
        #Incluye filtro de aire acondicionado
        data = next((item for item in atrributes if item["id"] == "INCLUDES_AC_FILTER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye filtro de aire acondicionado'] = data["value_name"]
        #Incluye filtro de combustible
        data = next((item for item in atrributes if item["id"] == "INCLUDES_FUEL_FILTER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye filtro de combustible'] = data["value_name"]
        #Incluye aceite
        data = next((item for item in atrributes if item["id"] == "INCLUDES_OIL"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Incluye aceite'] = data["value_name"]
        #Tipo de filtro de aceite
        data = next((item for item in atrributes if item["id"] == "OIL_FILTER_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de filtro de aceite'] = data["value_name"]
        #Formato
        data = next((item for item in atrributes if item["id"] == "FORMAT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Formato'] = data["value_name"]
        #Es apto para lavarropas
        data = next((item for item in atrributes if item["id"] == "IS_SUITABLE_FOR_WASHING_MACHINE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Es apto para lavarropas'] = data["value_name"]
        #Es blanqueador
        data = next((item for item in atrributes if item["id"] == "IS_BLEACHER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Es blanqueador'] = data["value_name"]
        #Tipo de funcionamiento de la válvula EGR
        data = next((item for item in atrributes if item["id"] == "EGR_VALVE_OPERATION_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de funcionamiento de la válvula EGR'] = data["value_name"]
        #Género del terminal
        data = next((item for item in atrributes if item["id"] == "TERMINAL_GENDER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Género del terminal'] = data["value_name"]
        #Diámetro de los puertos de vacío
        data = next((item for item in atrributes if item["id"] == "VACUUM_PORT_DIAMETER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Diámetro de los puertos de vacío'] = data["value_name"]
        #Cantidad de puertos de vacío
        data = next((item for item in atrributes if item["id"] == "VACUUM_PORT_QUANTITY"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de puertos de vacío'] = data["value_name"]
        #Cantidad de terminales
        data = next((item for item in atrributes if item["id"] == "TERMINAL_QUANTITY"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de terminales'] = data["value_name"]
        #Cantidad de huecos de montaje de la válvula EGR
        data = next((item for item in atrributes if item["id"] == "EGR_VALVE_MOUNTING_HOLE_QUANTITY"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de huecos de montaje de la válvula EGR'] = data["value_name"]
        #Juntas incluidas
        data = next((item for item in atrributes if item["id"] == "GASKETS_INCLUDED"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Juntas incluidas'] = data["value_name"]
        #Posición de la rótula de suspensión
        data = next((item for item in atrributes if item["id"] == "SUSPENSION_BALL_JOINT_POSITION"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Posición de la rótula de suspensión'] = data["value_name"]
        #Tuerca incluida
        data = next((item for item in atrributes if item["id"] == "CASTLE_NUT_INCLUDED"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tuerca incluida'] = data["value_name"]
        #Diámetro de entrada
        data = next((item for item in atrributes if item["id"] == "INLET_DIAMETER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Diámetro de entrada'] = data["value_name"]
        #Diámetro de salida
        data = next((item for item in atrributes if item["id"] == "OUTLET_DIAMETER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Diámetro de salida'] = data["value_name"]
        #Número de DOT
        data = next((item for item in atrributes if item["id"] == "DOT_NUMBER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Número de DOT'] = data["value_name"]
        #Grado del aceite de motor
        data = next((item for item in atrributes if item["id"] == "ENGINE_OIL_GRADE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Grado del aceite de motor'] = data["value_name"]
        #Tipo de aceite de motor
        data = next((item for item in atrributes if item["id"] == "ENGINE_OIL_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de aceite de motor'] = data["value_name"]
        #Categoría de servicio
        data = next((item for item in atrributes if item["id"] == "SERVICE_CATEGORY"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Categoría de servicio'] = data["value_name"]
        #Volumen del aceite de motor
        data = next((item for item in atrributes if item["id"] == "ENGINE_OIL_VOLUME"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Volumen del aceite de motor'] = data["value_name"]
        #Tipo de contenedor
        data = next((item for item in atrributes if item["id"] == "CONTAINER_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de contenedor'] = data["value_name"]
        #Ubicación
        data = next((item for item in atrributes if item["id"] == "BRAKE_LIGHT_POSITION"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Ubicación'] = data["value_name"]
        #Tipo de luz
        data = next((item for item in atrributes if item["id"] == "BULBS_TYPE"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Tipo de luz'] = data["value_name"]
        #Peso de la unidad
        data = next((item for item in atrributes if item["id"] == "UNIT_WEIGHT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Peso de la unidad'] = data["value_name"]
        #Formato del abrillantador
        data = next((item for item in atrributes if item["id"] == "BRIGHTENER_FORMAT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Formato del abrillantador'] = data["value_name"]
        #Cantidad de unidades
        data = next((item for item in atrributes if item["id"] == "UNITS_NUMBER"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Cantidad de unidades'] = data["value_name"]
        #Contenido neto
        data = next((item for item in atrributes if item["id"] == "NET_CONTENT"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Contenido neto'] = data["value_name"]
        #Set
        data = next((item for item in atrributes if item["id"] == "SET"), None)
        if data is not None :
            if data["value_name"] is not None:
                dataitem['Set'] = data["value_name"]
        #Por si se añade otro valor
        # data = next((item for item in atrributes if item["id"] == ""), None)
        # if data is not None :
        #     if data["value_name"] is not None:
        #         dataitem[''] = data["value_name"]
        
        print(json.dumps(dataitem, indent = 4,ensure_ascii=False))
        return(dataitem)
    except Exception as excep:
        print(excep)

def makeExcelLetters():
    abecedario = list(string.ascii_uppercase)
    contador = 1
    abecedarioExcel = []
    abecedario = list(string.ascii_uppercase)
    contador = 1
    abecedarioExcel = []
    while contador < 183:
        if contador <= 26:
            abecedarioExcel.append(abecedario[contador-1])
        elif(contador >= 27 and contador <= 52):
            abecedarioExcel.append('A'+abecedario[contador-27])
        elif(contador >= 53 and contador <= 78):
            abecedarioExcel.append('B'+abecedario[contador-53])
        elif(contador >= 79 and contador <= 104):
            abecedarioExcel.append('C'+abecedario[contador-79])
        elif(contador >= 105 and contador <= 130):
            abecedarioExcel.append('D'+abecedario[contador-105])
        elif(contador >= 131 and contador <= 156):
            abecedarioExcel.append('E'+abecedario[contador-131])
        elif(contador >= 157 and contador <= 182):
            abecedarioExcel.append('F'+abecedario[contador-157])
        contador += 1
    return abecedarioExcel
        
def makeexcel(items):
    try:
        abecedario = makeExcelLetters()
        #Colores
        #A1 a T1  (1 a 20) AE1 A AH1 (31 A 34)
        parteAmarilla = NamedStyle(name="parteAmarilla")
        parteAmarilla.font = Font(bold=True, color="FFFFFF")
        parteAmarilla.fill = PatternFill("solid", fgColor="FF9900")
        #U1 A AD1 (21 A 30)
        parteAzul = NamedStyle(name="parteAzul")
        parteAzul.font = Font(bold=True, color="FFFFFF")
        parteAzul.fill = PatternFill("solid", fgColor="0066CC")

        #AI1 A FR1 (35 A 174)
        parteVerde = NamedStyle(name="parteVerde")
        parteVerde.font = Font(bold=True, color="FFFFFF")
        parteVerde.fill = PatternFill("solid", fgColor="034A03")
        #Columna C 
        TituloNegrita = NamedStyle(name="TituloNegrita")
        TituloNegrita.font = Font(bold=True)

        wb = Workbook()
        hoja = wb.active
        wb.add_named_style(parteAmarilla)
        wb.add_named_style(parteAzul)
        wb.add_named_style(parteVerde)
        wb.add_named_style(TituloNegrita)
        hoja.append(listTitlesExcel)

        for i,data in enumerate(listTitlesExcel):
            if ((i+1) < 21) or ((i+1) >= 31 and (i+1) <= 34):
                hoja[abecedario[i]+"1"].style = 'parteAmarilla'
            elif((i+1)>= 21 and (i+1) <= 30):
                hoja[abecedario[i]+"1"].style = 'parteAzul'
            else:
                hoja[abecedario[i]+"1"].style = 'parteVerde'

        if items is not None:
            for contador,changeitem in enumerate(items):
                dataToAdd = []
                if changeitem is not None:
                    for i,data in enumerate(listRowExcel):
                        dataToAdd.append(changeitem.get(data,""))
                    hoja.append(dataToAdd)
                    hoja["C"+str(contador+2)].style = 'TituloNegrita'

        actualDate = datetime.today().strftime('%d-%m-%Y-%H-%M-%S')
        nameFile ='productos'+actualDate+'.xlsx'
        wb.save(nameFile)
        with open(nameFile, 'rb') as excel:
            created_at = timezone.now()
            updated_at = created_at
            data = {
                    "document" : File(excel),
                    "description": "Autogenerado con fecha "+actualDate,
                    "created_at" : created_at,
                    "updated_at" : updated_at
                }
            #save the document
            DocumentItems(**data).save()
        if os.path.exists(nameFile):
            os.remove(nameFile)
        else:
            print("The file does not exist")
    except Exception as excep:
        print(excep)

def readexcel():
    excelRaw = pandas.read_excel('Lista total de publicaciones CI - Para cesar Santana (pruebas).xlsx')
    excelDict = excelRaw.to_dict('records')
    saveCount = 0
    notSaveCount = 0
    error = 0
    for item in excelDict:
        try:
            olditem = DictionaryItems.objects.get(idMercadoLibre=item['Id'])
            notSaveCount += 1
        except DictionaryItems.DoesNotExist:
            number_part = str(item['SKU'])
            model = str(item['Atributo_x000D_\nModelo'])
            long_brand = str(item['Atributo_x000D_\nMarca'])
            short_brand = long_brand[:4]
            if (number_part == "nan"):
                if (str(item['Atributo_x000D_\nNúmero de parte']) != "nan"):
                    number_part = str(item['Atributo_x000D_\nNúmero de parte'])
                else:
                    number_part = None
            if (model == "nan"):
                model = None
            if (long_brand == "nan"):
                long_brand = "No registrado"
            if (short_brand == "nan"):
                short_brand = "NO R"
            data = {
                "idMercadoLibre":item['Id'],
                "long_brand":long_brand,
                "short_brand":short_brand,
                "number_part":number_part,
                "model":model,
                "stock":str(item['Stock']),
                "price":str(item['Precio'])
            }
            DictionaryItems(**data).save()
            saveCount += 1
        except Exception as excep:
            error += 1
    results = "Se agregaron: "+str(saveCount)+" Se omitieron: "+str(notSaveCount)+" Hubo error en:"+str(error)
    print (results)
    return (results)
